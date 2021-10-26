from openpyxl import Workbook
from random import *

wb = Workbook()
ws = wb.active
ws.title = "NadoSheet"

ws["A1"] = 1
ws["A2"] = 2
ws["A3"] = 3

ws["B1"] = 4
ws["B2"] = 5
ws["B3"] = 6

print(ws["A1"])
print(ws["A1"].value)
print(ws["A10"].value)

print(ws.cell(column=1, row=1).value)
print(ws.cell(column=2, row=1).value)

c = ws.cell(column=3, row=1, value=10)
print(c.value)

index = 1

for i in range(1, 11):
    for j in range(1, 11):
        a = ws.cell(i, j, index)
        index += 1

wb.save("sample.xlsx")
