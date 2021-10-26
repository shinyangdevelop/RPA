from openpyxl import Workbook

wb = Workbook()
ws = wb.create_sheet()
ws.title = "MySheet"
ws.sheet_properties.tabColor = "8888FF"

ws1 = wb.create_sheet("YourSheet")
ws2 = wb.create_sheet("NewSheet", 2)

new_ws = wb["NewSheet"]

print(wb.sheetnames)

new_ws["A1"] = "Test"
target = wb.copy_worksheet(new_ws)
target.title = "Copied Sheet"

wb.save("sample.xlsx")