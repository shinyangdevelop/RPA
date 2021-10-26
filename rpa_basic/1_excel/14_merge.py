from openpyxl import Workbook
wb = Workbook()
ws = wb.active

ws.merge_cells('B2:D2')
ws["B2"] = "Merged Cell"

wb.save("sample_merge.xlsx")