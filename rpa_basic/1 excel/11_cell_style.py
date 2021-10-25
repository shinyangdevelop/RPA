from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment

wb = load_workbook("sample.xlsx")
ws = wb.active

a1 = ws.cell(1, 1)
b1 = ws.cell(1, 2)
c1 = ws.cell(1, 3)

ws.column_dimensions["A"].width = 5

ws.row_dimensions[1].height = 50

a1.font = Font(color="FF8888", italic=True, bold=True)
b1.font = Font(color="88FF88", name="Arial", strike=True)
c1.font = Font(color="8888FF", size=20, underline="single")

thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"),
                     bottom=Side(style="thin"))

a1.border = thin_border
b1.border = thin_border
c1.border = thin_border

for row in ws.rows:
    for cell in row:
        cell.alignment = Alignment(horizontal="center", vertical="center")

        if cell.column == 1:
            continue
        if isinstance(cell.value, int) and cell.value > 90:
            cell.fill = PatternFill(fgColor="88FF88", fill_type="solid")
            cell.font = Font(color="8888FF")

ws.freeze_panes = "B2"

wb.save("sample_style.xlsx")
